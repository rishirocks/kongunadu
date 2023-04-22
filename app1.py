from flask import Flask,render_template,request,session
from openpyxl import load_workbook
from werkzeug.utils import secure_filename


app = Flask(__name__)
app.secret_key = 'mysecretkey'
@app.route("/",methods=["POST","GET"])

def home():
    reg_no = ""
    empty_row1 = None
    empty_row2 = None
    empty_row3 = None
    empty_row4 = None
    empty_row5 = None
    empty_row6 = None
    empty_row7 = None
    empty_row8 = None

    if request.method == "POST":
         if 'file' in request.files:
             file = request.files['file']
             filename = secure_filename(file.filename)
             file.save(filename)
             session['filename']=filename

         if request.form.get('Reg_no'):
            Reg_no = request.form["Reg_no"]
            reg_no = '6213' + Reg_no
         else:
            Reg_no = None

         if request.form.get('Name'):
            Name = request.form["Name"]
            Names = session.get('Names',[])
            Names.append(Name)
            session['Names'] = Names
         else:
            Names = session.get('Names',[])
            Name = None

         if request.form.get('IA_1'):
            IA_1 = int(request.form['IA_1'])
            IA_one = float((IA_1 * 30) / 100)
            IA_1s = session.get('IA_1s',[])
            IA_1s.append(IA_1)
            session['IA_1s'] = IA_1s
         else:
            IA_1s =  session.get('IA_1s',[])
            IA_1   = None
            IA_one = None

         if request.form.get('Assi'):
             Assi = int(request.form['Assi'])
             A1 = session.get('A1',[])
             A1.append(Assi)
             session['A1'] = A1
         else:
             A1 = session.get('A1',[])
             Assi = None

         if request.form.get('IA_2'):
             IA_2 = int(request.form['IA_2'])
             IA_two = float((IA_2 * 30) / 100)
             IA_2s = session.get('IA_2s',[])
             IA_2s.append(IA_2)
             session['IA_2s'] =IA_2s
         else:
             IA_2s = session.get('IA_2s',[])
             IA_2   = None
             IA_two = None

         if request.form.get('Assign'):
             Assign = int(request.form.get('Assign'))
             A2 = session.get('A2', [])
             A2.append(Assign)
             session['A2'] = A2
         else:
             A2 = session.get('A2',[])
             Assign = None

         if request.form.get('IA_3'):
             IA_3 = int(request.form['IA_3'])
             IA_three = float((IA_3 * 60)/100)
             IA_3s = session.get('IA_3s', [])
             IA_3s.append(IA_3)
             session['IA_3s'] = IA_3s
         else:
             IA_3s = session.get('IA_3s',[])
             IA_3 = None
             IA_three = None

         filename = session['filename']
         wb = load_workbook(filename)
         sheet = wb.active

         if sheet.cell(row=1, column=1).value is None:
            sheet.cell(row=1,column=1,value='Reg_no')
            sheet.cell(row=1,column=2,value='Name')
            sheet.cell(row=1, column=3, value='IA-1(100)')
            sheet.cell(row=1, column=4, value='IA-1(30)')
         if request.form.get('Assi') and sheet.cell(row=1,column=5).value is None:
            sheet.cell(row=1, column=5, value='Ass-1')
         if request.form.get('IA_2') and sheet.cell(row=1,column=6).value is None:
            sheet.cell(row=1, column=6, value='IA-2(100)')
            sheet.cell(row=1, column=7, value='IA-2(30)')
         if request.form.get('Assign') and sheet.cell(row=1,column=8).value is None:
            sheet.cell(row=1, column=8, value='Ass-2')
         if request.form.get('IA_3') and sheet.cell(row=1,column=9).value is None:
            sheet.cell(row=1, column=9, value='IA_3(100)')
            sheet.cell(row=1, column=10, value='IA_3(60)')

         for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row,column=1).value == Reg_no:
               if request.form.get('IA_1'):
                  sheet.cell(row=row,column=3).value = IA_1
                  sheet.cell(row=row,column=4).value = IA_one
               if request.form.get('Assi'):
                  sheet.cell(row=row,column=5).value = Assi
               if request.form.get('IA_2'):
                  sheet.cell(row=row, column=6).value = IA_2
                  sheet.cell(row=row, column=7).value = IA_two
               if request.form.get('Assign'):
                  sheet.cell(row=row, column=8).value = Assign
               if request.form.get('IA_3'):
                   sheet.cell(row=row, column=9).value = IA_3
                   sheet.cell(row=row, column=10).value = IA_three
               break
            elif request.form.get('IA_1'):
               if sheet.cell(row=row,column=3).value is None and sheet.cell(row=row,column=4).value is None:
                  empty_row1 = row
                  empty_row2 = row
                  break
            elif request.form.get('Assi'):
                if sheet.cell(row=row,column=5).value is None:
                   empty_row3 = row
                   break
            elif request.form.get('IA_2'):
                if sheet.cell(row=row, column=6).value is None and sheet.cell(row=row, column=7).value is None:
                   empty_row4 = row
                   empty_row5 = row
                   break
            elif request.form.get('Assign'):
                if sheet.cell(row=row,column=8).value is None:
                   empty_row6 = row
                   break
            elif request.form.get('IA_3'):
                if sheet.cell(row=row, column=9).value is None and sheet.cell(row=row, column=10).value is None:
                   empty_row7 = row
                   empty_row8 = row
                   break
         else:
             data = [reg_no,Name,IA_1,IA_one,Assi,IA_2,IA_two,Assign,IA_3,IA_three]
             sheet.append(data)
         if empty_row1 is not None and empty_row2 is not None:
             sheet.cell(row=empty_row1, column=2).value = IA_1
             sheet.cell(row=empty_row2, column=3).value = IA_one
         if  empty_row3 is not None:
            sheet.cell(row=empty_row3, column=4).value = Assi
         if empty_row4 is not None and empty_row5 is not None:
            sheet.cell(row=empty_row4, column=5).value = IA_2
            sheet.cell(row=empty_row5, column=6).value = IA_two
         if empty_row6 is not None:
            sheet.cell(row=empty_row6, column=7).value = Assign
         if empty_row7 is not None and empty_row8 is not None:
             sheet.cell(row=empty_row7, column=8).value = IA_3
             sheet.cell(row=empty_row8, column=9).value = IA_three

         wb.save(filename)
         return render_template("home.html", choose_file=False,Names=Names,IA_1s=IA_1s,IA_one=IA_one,A1=A1,IA_2s=IA_2s,IA_two=IA_two,A2=A2,IA_3s=IA_3s,IA_three=IA_three)
    return render_template("home.html", choose_file=True)

if __name__=="__main__":
    app.run(debug=True)

